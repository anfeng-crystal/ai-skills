import io
from typing import List, Dict, Any

import pandas as pd
import docx2txt
from openpyxl import load_workbook


def read_docx(file_bytes: bytes) -> str:
    """Extract plain text from a .docx file."""
    with io.BytesIO(file_bytes) as f:
        return docx2txt.process(f)


def read_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Convert each sheet of a .xlsx to a list of row dictionaries.
    Returns a list like [{"sheet": "Sheet1", "data": [{col: val, ...}, ...]}, ...]
    """
    with io.BytesIO(file_bytes) as f:
        wb = load_workbook(f, data_only=True, read_only=True)
        sheets: List[Dict[str, Any]] = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = rows[0]
            data = [
                {str(header[i]): row[i] for i in range(len(header))}
                for row in rows[1:]
                if any(cell is not None for cell in row)
            ]
            sheets.append({"sheet": ws.title, "data": data})
        return sheets
