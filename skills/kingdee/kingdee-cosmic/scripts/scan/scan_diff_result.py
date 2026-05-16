# -*- coding: utf-8 -*-
"""
代码评审结果差异比对工具
功能：
1. 从命令行参数获取检测JSON结果
2. 读取 sonarfile.xlsx 数据
3. 按照「规则编码 + 违规文件路径」组合，判断差异
4. 对比count字段，判断违规数量是否变化
5. 返回差异结果（新增违规、已修复违规、count变化违规、未变化违规）
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误: 请先安装 openpyxl 库")
    print("执行命令: pip install openpyxl")
    exit(1)

# 技能根目录（当前脚本所在目录）
SCRIPT_DIR = Path(__file__).parent

# 项目根目录（相对于脚本位置的固定路径）
# 脚本位于: scripts/ 目录，项目根目录是脚本的父目录的父目录
PROJECT_ROOT = SCRIPT_DIR.parent.parent

# sonarfile.xlsx 默认路径
DEFAULT_SONAR_FILE = SCRIPT_DIR.parent / "references" / "SonarFile.xlsx"


def load_json_result(json_input):
    """
    加载JSON检测结果
    
    参数:
        json_input: JSON字符串或JSON文件路径
    
    返回:
        list: 违规列表
    """
    # 判断是文件路径还是JSON字符串
    if isinstance(json_input, str):
        if os.path.exists(json_input):
            # 是文件路径
            with open(json_input, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            # 尝试解析为JSON字符串
            try:
                return json.loads(json_input)
            except json.JSONDecodeError as e:
                print(f"错误: 无法解析JSON输入: {e}")
                return []
    elif isinstance(json_input, list):
        return json_input
    else:
        print(f"错误: 不支持的输入类型 {type(json_input)}")
        return []


def load_sonar_excel(excel_path):
    """
    读取 sonarfile.xlsx 数据
    
    参数:
        excel_path: Excel文件路径
    
    返回:
        list: 历史违规列表，每项包含 rule_code 和 file_path
    """
    if not os.path.exists(excel_path):
        print(f"警告: Excel文件不存在: {excel_path}")
        return []
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        
        # 获取第一个工作表
        sheet = workbook.active
        if sheet is None:
            print(f"错误: Excel文件没有工作表")
            return []
        
        # 读取表头，找到列索引
        headers = []
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for cell_value in header_row:
            headers.append(str(cell_value).strip() if cell_value else "")
        
        # 查找关键列
        rule_code_col = None
        file_path_col = None
        rule_name_col = None
        rule_level_col = None
        line_number_col = None
        count_col = None
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if '规则编码' in header or 'rule_code' in header_lower:
                rule_code_col = idx
            elif '所在文件' in header or 'file_path' in header_lower:
                file_path_col = idx
            elif '漏洞名称' in header or '规则名称' in header or 'rule_name' in header_lower:
                rule_name_col = idx
            elif '等级' in header or 'rule_level' in header_lower:
                rule_level_col = idx
            elif '所在行数' in header or '行号' in header or 'line' in header_lower:
                line_number_col = idx
            elif '漏洞数量' in header or 'count' in header_lower:
                count_col = idx
        
        if rule_code_col is None or file_path_col is None:
            print(f"错误: Excel缺少必要列（规则编码、违规文件路径）")
            print(f"  表头: {headers}")
            return []
        
        # 读取数据行
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[rule_code_col] is None or row[file_path_col] is None:
                continue
            
            record = {
                'rule_code': str(row[rule_code_col]).strip(),
                'file_path': str(row[file_path_col]).strip(),
            }
            
            if rule_name_col is not None and len(row) > rule_name_col and row[rule_name_col]:
                record['rule_name'] = str(row[rule_name_col]).strip()
            if rule_level_col is not None and len(row) > rule_level_col and row[rule_level_col]:
                record['rule_level'] = str(row[rule_level_col]).strip()
            if line_number_col is not None and len(row) > line_number_col and row[line_number_col]:
                record['line_number'] = str(row[line_number_col]).strip()
            if count_col is not None and len(row) > count_col and row[count_col]:
                record['count'] = row[count_col]
            
            records.append(record)
        
        workbook.close()
        return records
        
    except Exception as e:
        print(f"错误: 读取Excel文件失败: {e}")
        return []


def build_key_map(records):
    """
    构建「规则编码 + 违规文件路径」组合的映射表
    
    参数:
        records: 违规记录列表
    
    返回:
        dict: key -> record 的映射表，key为(rule_code, file_path)元组
        同一文件同一规则的多条记录，count会累加
    """
    key_map = {}
    for record in records:
        rule_code = record.get('rule_code', '')
        file_path = record.get('file_path', '')
        line_number = record.get('line_number', 0)
        record_count = record.get('count', 1)
        
        if rule_code and file_path:
            # 标准化文件路径（统一使用正斜杠）
            normalized_path = file_path.replace('\\', '/')
            # 按 (rule_code, file_path) 分组，用于差异比对
            key = (rule_code, normalized_path)
            
            if key in key_map:
                # 已存在，累加 count
                key_map[key]['count'] = key_map[key].get('count', 1) + record_count
            else:
                # 新增记录，初始化 count
                new_record = record.copy()
                new_record['count'] = record_count if record_count else 1
                key_map[key] = new_record
    return key_map


def compare_results(current_records, history_records):
    """
    比对当前检测结果与历史记录，找出差异
    
    参数:
        current_records: 当前检测的违规列表
        history_records: 历史记录的违规列表
    
    返回:
        dict: 包含新增违规、已修复违规、count变化违规、未变化违规
    """
    # 构建映射表
    current_map = build_key_map(current_records)
    history_map = build_key_map(history_records)
    
    current_keys = set(current_map.keys())
    history_keys = set(history_map.keys())
    
    # 新增违规：当前有，历史没有
    new_violations = current_keys - history_keys
    
    # 已修复违规：历史有，当前没有
    fixed_violations = history_keys - current_keys
    
    # 两者都有的违规：需要对比count
    common_keys = current_keys & history_keys
    
    # 分类：count变化 vs count未变化
    count_changed_violations = set()
    unchanged_violations = set()
    count_changed_details = []
    
    for key in common_keys:
        current_record = current_map.get(key, {})
        history_record = history_map.get(key, {})
        
        # 获取count值
        current_count = current_record.get('count', 0)
        history_count = history_record.get('count', 0)
        
        # 尝试转换为整数进行比较
        try:
            current_count = int(current_count) if current_count else 0
        except (ValueError, TypeError):
            current_count = 0
        
        try:
            history_count = int(history_count) if history_count else 0
        except (ValueError, TypeError):
            history_count = 0
        
        if current_count != history_count:
            # count发生变化
            count_changed_violations.add(key)
            # 从当前记录或历史记录中获取规则名称
            rule_name = current_record.get('rule_name', '') or history_record.get('rule_name', '')
            rule_level = current_record.get('rule_level', '') or history_record.get('rule_level', '')
            count_changed_details.append({
                'rule_code': key[0],
                'rule_name': rule_name,
                'rule_level': rule_level,
                'file_path': key[1],
                'current_count': current_count,
                'history_count': history_count,
                'count_diff': current_count - history_count
            })
        else:
            # count未变化
            unchanged_violations.add(key)
    
    # 构建详细的新增违规列表（包含规则名称）
    new_violation_details = []
    new_violations_with_name = []
    for key in new_violations:
        if key in current_map:
            record = current_map[key]
            new_violation_details.append(record)
            rule_name = record.get('rule_name', '')
            new_violations_with_name.append((key[0], key[1], rule_name))
    
    # 构建详细的已修复违规列表（包含规则名称）
    fixed_violation_details = []
    fixed_violations_with_name = []
    for key in fixed_violations:
        if key in history_map:
            record = history_map[key]
            fixed_violation_details.append(record)
            rule_name = record.get('rule_name', '')
            fixed_violations_with_name.append((key[0], key[1], rule_name))
    
    return {
        'new_violations': new_violations_with_name,
        'fixed_violations': fixed_violations_with_name,
        'count_changed_violations': list(count_changed_violations),
        'unchanged_violations': list(unchanged_violations),
        'new_violation_details': new_violation_details,
        'fixed_violation_details': fixed_violation_details,
        'count_changed_details': count_changed_details,
        'stats': {
            'new_count': len(new_violations),
            'fixed_count': len(fixed_violations),
            'count_changed_count': len(count_changed_violations),
            'unchanged_count': len(unchanged_violations),
            'current_total': len(current_keys),
            'history_total': len(history_keys)
        }
    }


def generate_diff_report(diff_result, output_format='json'):
    """
    生成差异报告 - 输出新增违规、已修复违规、Count变化违规
    
    参数:
        diff_result: compare_results 返回的结果
        output_format: 输出格式，'json' 或 'markdown'
    
    返回:
        str: 格式化的报告
    """
    if output_format == 'json':
        return json.dumps(diff_result, ensure_ascii=False, indent=2)

    elif output_format == 'markdown':
        lines = []
        lines.append("# 代码评审差异报告")
        lines.append("")
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        stats = diff_result['stats']
        
        # 1. 新增违规：当前有，历史没有（完全新增）
        if diff_result['new_violation_details']:
            lines.append("## 新增违规")
            lines.append("")
            lines.append(f"共 {stats['new_count']} 条新增违规")
            lines.append("")
            lines.append("| 规则编码 | 规则名称 | 风险等级 | 违规文件路径 | 数量 |")
            lines.append("|----------|----------|----------|--------------|------|")
            for detail in sorted(diff_result['new_violation_details'], key=lambda x: (x.get('rule_code', ''), x.get('file_path', ''))):
                rule_code = detail.get('rule_code', '')
                rule_name = detail.get('rule_name', '')
                rule_level = detail.get('rule_level', '')
                file_path = detail.get('file_path', '')
                count = detail.get('count', 1)
                lines.append(f"| {rule_code} | {rule_name} | {rule_level} | {file_path} | {count} |")
            lines.append("")
        
        # 2. 已修复违规：历史有，当前没有（完全修复）
        if diff_result['fixed_violation_details']:
            lines.append("## 已修复违规")
            lines.append("")
            lines.append(f"共 {stats['fixed_count']} 条已修复违规")
            lines.append("")
            lines.append("| 规则编码 | 规则名称 | 风险等级 | 违规文件路径 | 数量 |")
            lines.append("|----------|----------|----------|--------------|------|")
            for detail in sorted(diff_result['fixed_violation_details'], key=lambda x: (x.get('rule_code', ''), x.get('file_path', ''))):
                rule_code = detail.get('rule_code', '')
                rule_name = detail.get('rule_name', '')
                rule_level = detail.get('rule_level', '')
                file_path = detail.get('file_path', '')
                count = detail.get('count', 1)
                lines.append(f"| {rule_code} | {rule_name} | {rule_level} | {file_path} | {count} |")
            lines.append("")
        
        # 3. Count 变化违规：rule_code+file_path 相同，但 count 变化
        if diff_result['count_changed_details']:
            lines.append("## Count 变化违规")
            lines.append("")
            lines.append(f"共 {stats['count_changed_count']} 条违规数量发生变化")
            lines.append("")
            lines.append("| 规则编码 | 规则名称 | 风险等级 | 违规文件路径 | 历史数量 | 当前数量 | 变化量 |")
            lines.append("|----------|----------|----------|--------------|----------|----------|--------|")
            for detail in sorted(diff_result['count_changed_details'], key=lambda x: (x['rule_code'], x['file_path'])):
                rule_name = detail.get('rule_name', '')
                rule_level = detail.get('rule_level', '')
                count_diff = detail['count_diff']
                diff_str = f"+{count_diff}" if count_diff > 0 else str(count_diff)
                lines.append(f"| {detail['rule_code']} | {rule_name} | {rule_level} | {detail['file_path']} | {detail['history_count']} | {detail['current_count']} | {diff_str} |")
            lines.append("")
        
        # 4. 无任何变化
        if not diff_result['new_violation_details'] and not diff_result['fixed_violation_details'] and not diff_result['count_changed_details']:
            lines.append("## 无变化")
            lines.append("")
            lines.append("没有检测到任何变化。")
            lines.append("")
        
        return '\n'.join(lines)

    else:
        return json.dumps(diff_result, ensure_ascii=False, indent=2)


def main():
    """
    主函数
    """
    parser = argparse.ArgumentParser(description='代码评审结果差异比对工具')
    parser.add_argument('--json', dest='json_input', 
                        help='JSON检测结果（文件路径或JSON字符串）')
    parser.add_argument('--excel', dest='excel_path', 
                        default=str(DEFAULT_SONAR_FILE),
                        help=f'sonarfile.xlsx文件路径（默认: {DEFAULT_SONAR_FILE}）')
    parser.add_argument('--format', dest='output_format', 
                        choices=['json', 'markdown'], default='markdown',
                        help='输出格式（json或markdown，默认markdown）')
    parser.add_argument('--output', dest='output_file',
                        help='输出文件路径（不指定则输出到控制台）')
    
    args = parser.parse_args()
    
    # 1. 加载当前检测结果
    if args.json_input:
        current_records = load_json_result(args.json_input)
    else:
        # 从标准输入读取JSON
        print("请输入JSON检测结果（按Ctrl+D结束）:")
        try:
            json_input = sys.stdin.read()
            current_records = load_json_result(json_input)
        except KeyboardInterrupt:
            print("\n操作已取消")
            return
    
    if not current_records:
        print("错误: 未获取到有效的检测结果")
        return
    
    print(f"[信息] 当前检测结果: {len(current_records)} 条记录")
    
    # 2. 加载历史记录
    history_records = load_sonar_excel(args.excel_path)
    print(f"[信息] 历史记录: {len(history_records)} 条记录")
    
    # 3. 比对差异
    diff_result = compare_results(current_records, history_records)
    
    # 4. 生成报告
    report = generate_diff_report(diff_result, args.output_format)
    
    # 5. 输出结果到固定路径
    if args.output_file:
        output_path = Path(args.output_file)
    else:
        # 使用固定输出路径
        output_path = SCRIPT_DIR.parent / "result" / "scan_java_diff.md"
    
    # 确保目录存在
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 写入文件
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"[信息] 报告已保存到: {output_path}")
    
    return diff_result


if __name__ == '__main__':
    main()
